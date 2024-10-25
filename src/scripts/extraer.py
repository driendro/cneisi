import openpyxl
from django.db.models import Q
from core.models import UserAsistente, Dependencia, TallesRemeras


def export_user_asistentes_to_excel():
    """Exports UserAsistente data to an Excel file."""

    # Create a new workbook and worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Set column headers
    headers = [
        'Apellido',
        'Nombre',
        'Documento',
        'CUIT',
        'Fecha de Nacimiento',
        'Teléfono Personal',
        'Teléfono Emergencia',
        'Dependencia',
        'Grupo Sanguíneo',
        'Régimen de Comida',
        'Régimen de Comida Otro',
        'Talle de Ropa',
        'Alergia',
        'Alergia Otro',
        'Medicamento',
        'Medicamento Otro',
    ]
    for col, header in enumerate(headers, 1):
        worksheet.cell(row=1, column=col, value=header)

    # Fetch UserAsistente data
    user_asistentes = UserAsistente.objects.all()

    # Populate worksheet with data
    row = 2
    for asistente in user_asistentes:
        worksheet.cell(row=row, column=1, value=asistente.user.last_name.upper())
        worksheet.cell(row=row, column=2, value=asistente.user.first_name.title())
        worksheet.cell(row=row, column=3, value=asistente.documento)
        worksheet.cell(row=row, column=4, value=asistente.cuit)
        worksheet.cell(row=row, column=5, value=asistente.fecha_nacimiento.strftime('%d/%m/%Y'))
        worksheet.cell(row=row, column=6, value=str(asistente.telefono_personal))
        worksheet.cell(row=row, column=7, value=str(asistente.telefono_emergencia))
        worksheet.cell(row=row, column=8, value=asistente.dependencia.nombre_corto)
        worksheet.cell(row=row, column=9, value=asistente.grupo_sangineo)
        worksheet.cell(row=row, column=10, value=asistente.regimen_comida)
        worksheet.cell(row=row, column=11, value=asistente.regimen_comida_otro)
        worksheet.cell(row=row, column=12, value=asistente.talle_ropa.designacion)
        worksheet.cell(row=row, column=13, value=asistente.alergia)
        worksheet.cell(row=row, column=14, value =asistente.alergia_otro)
        worksheet.cell(row=row, column=15, value =asistente.medicamento)
        worksheet.cell(row=row, column=16, value = asistente.medicamento_otro)
        row += 1

        # Save the Excel file
        workbook.save('/home/ronconi/user_asistentes.xlsx')
        print("Excel file exported successfully!")
